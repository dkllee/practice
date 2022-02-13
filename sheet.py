from openpyxl import Workbook
wb=Workbook()
ws=wb.create_sheet()
ws.title="mysheet"
ws.sheet_properties.tabColor="ff66ff"

ws1=wb.create_sheet("yoursheet")
ws2=wb.create_sheet("newsheet",1)

new_ws=wb["newsheet"]

print(wb.sheetnames)


#sheet 복사

new_ws["A1"]="test"
target=wb.copy_worksheet(new_ws)
target.title="copy sheet"

wb.save("bbbb.xlsx")
