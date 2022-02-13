from openpyxl import Workbook
from random import *
wb=Workbook()
ws=wb.active

# #1줄씩 데이터 넣기
# ws.append(["번호","영어","수학"])
# for i in range(1,11):
#     ws.append([i,randint(0,100),randint(0,100)])


# # col_B=ws["B"]
# # print(col_B)

# # for cell in col_B:
#     # print(cell.value)


# col_range=ws["B:C"]

# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

wb.save("ee.xlsx")




언제 온데??

