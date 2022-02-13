from openpyxl import Workbook

wb=Workbook()
ws=wb.active
ws.title="dddd"

ws["A1"]="장현경"
ws["A2"]="왜 웃어"
ws["A3"]="죽을래"
ws["A4"]="누구야"
ws["B1"]="남친 누구?"

# ws.cell(column=3,row=1,value=99) #ws["C1"]=99

# print(ws["A1"])
# print(ws["A1"].value)
# print(ws["A122"].value) #값이 없으면 None 으로 출력

# print(ws.cell(row=3, column=1).value) #row=행, column=열f
# print(ws.cell(row=4, column=1).value)


# from random import *
# index=1
for x in range(1,101):
     for y in range(1,101):
          ws.cell(row=x, column=y, value="남친 누구??")
#         ws.cell(row=x,column=y,value=index)
#         index+=1



wb.save("ee.xlsx")
